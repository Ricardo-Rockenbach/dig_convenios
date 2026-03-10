from regras import obter_regra
import pandas as pd
import os
from openpyxl import load_workbook
from auxiliar import selecionar_DADOS_BRUTOS
from auxiliar import selecionar_PLANILHA_CONVENIOS_20XX


def processar_planilha():
    # =========================
    # CONFIGURAÇÕES
    # =========================

    # Mapa dos Meses
    meses = {
        1: "JANEIRO",
        2: "FEVEREIRO",
        3: "MARÇO",
        4: "ABRIL",
        5: "MAIO",
        6: "JUNHO",
        7: "JULHO",
        8: "AGOSTO",
        9: "SETEMBRO",
        10: "OUTUBRO",
        11: "NOVEMBRO",
        12: "DEZEMBRO"
    }

    # Mapa das linhas dos meses na planilha a ser preenchida
    posicao_meses = {
        "JANEIRO": 3,
        "FEVEREIRO": 34,
        "MARÇO": 65,
        "ABRIL": 96,
        "MAIO": 127,
        "JUNHO": 158,
        "JULHO": 189,
        "AGOSTO": 220,
        "SETEMBRO": 251,
        "OUTUBRO": 282,
        "NOVEMBRO": 313,
        "DEZEMBRO": 344
    }

    def encontrar_linha_convenio(ws, posicao_mes, convenio):
        
        linha = posicao_mes

        while True:
            valor_celula = ws.cell(row=linha, column=1).value
            
            # se encontrou o convênio
            if str(valor_celula).strip() == str(convenio).strip():
                return linha

            # se chegou no próximo mês, parar
            if str(valor_celula).strip() == "TOTAL":
                return None

            linha += 1  

    def escrever_valores(ws, linha, dados):

        coluna_inicial = 6  # Coluna F

        ws.cell(row=linha, column=coluna_inicial).value = dados["VALOR_CUSTO_M"]
        ws.cell(row=linha, column=coluna_inicial + 1).value = dados["VALOR_VENDA_M"]

        ws.cell(row=linha, column=coluna_inicial + 2).value = dados["VALOR_CUSTO_D"]
        ws.cell(row=linha, column=coluna_inicial + 3).value = dados["VALOR_VENDA_D"]

        ws.cell(row=linha, column=coluna_inicial + 4).value = dados["VALOR_CUSTO_T"]
        ws.cell(row=linha, column=coluna_inicial + 5).value = dados["VALOR_VENDA_T"]

        if "SAUDE MAIOR" in dados.name:
            ws.cell(row=linha, column=coluna_inicial + 7).value = dados["TOTAL_VENDA"]


    # =========================
    # LEITURA DOS DADOS
    # =========================

    # Ler o arquivo xlsx
    caminho_arquivo = selecionar_DADOS_BRUTOS()

    df = pd.read_excel(
        caminho_arquivo,
        usecols=[
            "CONVENIO",
            "VALOR_CUSTO_M",
            "VALOR_VENDA_M",
            "VALOR_CUSTO_D",
            "VALOR_VENDA_D",
            "VALOR_CUSTO_T",
            "VALOR_VENDA_T", 
            "INDATA_INICIAL",
            "TOTAL_VENDA"
    ],
    )

    # Tratar a coluna "INDATA_INICIAL" para extrair o mês e o nome do mês
    df["INDATA_INICIAL"] = pd.to_datetime(df["INDATA_INICIAL"], errors="coerce", dayfirst=True)
    df["MES"] = df["INDATA_INICIAL"].dt.month
    df["MES_NOME"] = df["MES"].map(meses)


    # =========================
    # LIMPEZA DOS DADOS
    # =========================

    # remover linhas com valores nulos
    df = df.dropna(subset=["CONVENIO"])
    # definir index como a coluna "CONVENIO"
    df = df.set_index("CONVENIO")
    # Remover os covenios com nome "CONVENIO"
    df = df[~df.index.astype(str).str.contains("CONVENIO", case=False)]
    # dropar a coluna "INDATA_INICIAL"
    df = df.drop(columns=["INDATA_INICIAL"])


    # =========================
    # CRIAR PLANILHAS POR CONVÊNIO
    # =========================

    # criar a pasta "data/auditoria" se ela não existir
    os.makedirs("data/auditoria", exist_ok=True)
    

    # abrir planilha oficial
    caminho_arquivo = selecionar_PLANILHA_CONVENIOS_20XX()
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    for convenio, df_convenio in df.groupby(level=0):

        regra = obter_regra(convenio)

        for _, row in df_convenio.iterrows():

            row = regra.aplicar(row)

            mes_nome = row["MES_NOME"]

            linha_mes = posicao_meses[mes_nome]

            linha_convenio = encontrar_linha_convenio(ws, linha_mes, convenio)

            if linha_convenio is not None:

                escrever_valores(ws, linha_convenio, row)

            else:
                print(f"Convênio não encontrado: {convenio} em {mes_nome}")

    # criar uma nova planilha para cada convenio
    for convenio, df_convenio in df.groupby(level=0):
        print(f"Processando convênio: {convenio}")
        # ordenar o dataframe por mes
        df_convenio = df_convenio.sort_values("MES")
        df_convenio.to_excel(f"data\\auditoria\\{convenio}.xlsx")

    wb.save(caminho_arquivo)
    # fechar o arquivo para liberar o acesso
    wb.close()

